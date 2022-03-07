import yfinance as yf
from math import floor
from openpyxl import Workbook
import time
from os import path, mkdir

TOTAL_BALANCE = 10000
INVESTMENT_BALANCE = 5000
CASH_BALANCE = 5000

ASSETS = ['IWD', 'IEF', 'GLD', 'QQQ', 'SHY']
WEIGHTS = {'IWD': 0.25, 'IEF': 0.25, 'GLD': 0.25, 'QQQ': 0.25, 'SHY': 0.25} 

def num_purchase_shares():
    for asset in ASSETS:
        curr_price = yf.Ticker(asset).history(period="max")['Close'][-1]
        allocate_amount = INVESTMENT_BALANCE * WEIGHTS[asset]
        shares = floor(allocate_amount / curr_price)
        print(asset, curr_price, shares)

def report_to_excel():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = "Total Balance"   
    ws['A2'] = "Investment Balance"
    ws['A3'] = "Cash Balance"

    ws['B1'] = TOTAL_BALANCE
    ws['B2'] = INVESTMENT_BALANCE
    ws['B3'] = CASH_BALANCE

    if not path.exists('report'):    
        mkdir('./report')        
    
    filename = time.strftime("%Y-%m")
    wb.save("./report/{}.xlsx".format(filename))


if __name__ == "__main__":
    num_purchase_shares()
    report_to_excel()

